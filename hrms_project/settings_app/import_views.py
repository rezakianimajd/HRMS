"""Excel import views for the Settings module."""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from core.engines.import_engine import ImportEngine


def _get_company(request):
    return getattr(request, 'tenant', None) or getattr(request, 'company', None)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def import_types(request):
    """List all available import types."""
    return Response(ImportEngine.get_types())


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def import_template(request, import_type):
    """Download a blank Excel template for the given import type."""
    if import_type not in ImportEngine.IMPORT_TYPES:
        return Response({'error': 'نوع درون‌ریزی نامعتبر است'}, status=404)

    cfg = ImportEngine.IMPORT_TYPES[import_type]
    persian_headers = ImportEngine.get_persian_headers(import_type)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg['label']

    # Header row (Persian)
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    center = Alignment(horizontal='center', vertical='center')

    for col_idx, header in enumerate(persian_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(18, len(header) + 8)

    # Sample row
    for col_idx, sample_val in enumerate(cfg['sample'], start=1):
        cell = ws.cell(row=2, column=col_idx, value=sample_val)
        cell.alignment = center

    # Employee-specific: add dropdown data validations for choice columns
    if import_type == 'employees':
        from openpyxl.worksheet.datavalidation import DataValidation

        company = _get_company(request)
        lookups = ImportEngine.get_employee_choice_lists(company)

        # (headers index, lookup key) for columns requiring dropdowns
        dropdown_cols = {
            'gender': ['مرد', 'زن'],
            'marital_status': ['مجرد', 'متأهل', 'مطلقه', 'همسر فوت‌شده'],
            'status': ['شاغل', 'مرخصی طولانی‌مدت', 'بازنشسته', 'اخراج', 'فوت'],
            'work_shift': ['صبح', 'عصر', 'شیفتی', 'نامنظم'],
            'education_level': ['زیر دیپلم', 'دیپلم', 'کاردانی', 'کارشناسی', 'کارشناسی ارشد', 'دکتری'],
            'department': lookups.get('department', []),
            'job_title': lookups.get('job_title', []),
            'work_location': lookups.get('work_location', []),
            'insurance_list': lookups.get('insurance_list', []),
            'contract_type': lookups.get('contract_type', []),
        }

        # Map header key -> column index (1-based)
        col_idx_by_key = {key: idx for idx, key in enumerate(cfg['headers'], start=1)}

        for key, choices in dropdown_cols.items():
            if not choices:
                continue
            col_idx = col_idx_by_key.get(key)
            if not col_idx:
                continue
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            formula = '"' + ','.join(str(c) for c in choices) + '"'
            dv = DataValidation(
                type='list',
                formula1=formula,
                allow_blank=True,
                showDropDown=False,
            )
            dv.error = 'لطفاً از لیست انتخاب کنید'
            dv.errorTitle = 'مقدار نامعتبر'
            ws.add_data_validation(dv)
            # Apply to rows 2..1000
            dv.add(f'{col_letter}2:{col_letter}1000')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=template_{import_type}.xlsx'
    wb.save(response)
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_upload(request):
    """Upload an Excel file and import rows."""
    import_type = request.data.get('import_type') or request.query_params.get('import_type')
    if not import_type or import_type not in ImportEngine.IMPORT_TYPES:
        return Response({'error': 'نوع درون‌ریزی مشخص نشده است'}, status=400)

    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'فایل انتخاب نشده است'}, status=400)

    company = _get_company(request)

    try:
        rows = ImportEngine.parse_excel(file, import_type=import_type)
    except Exception as e:
        return Response({'error': f'خطا در خواندن فایل اکسل: {str(e)[:100]}'}, status=400)

    if not rows:
        return Response({'error': 'فایل اکسل خالی است'}, status=400)

    valid_rows, errors = ImportEngine.validate_rows(rows, import_type)

    if errors:
        return Response({
            'error': 'برخی ردیف‌ها خطا دارند',
            'validation_errors': errors,
            'total_rows': len(rows),
            'valid_rows': len(valid_rows),
            'imported_count': 0,
            'skipped': [],
        }, status=400)

    try:
        if import_type == 'employees':
            created, skipped = ImportEngine.import_employees(valid_rows, company)
        elif import_type in ('salary_records', 'salary_bulk'):
            year = request.data.get('year') or request.query_params.get('year')
            month = request.data.get('month') or request.query_params.get('month')
            year = int(year) if year else None
            month = str(month) if month else None
            created, skipped = ImportEngine.import_salaries(valid_rows, company, year=year, month=month)
        elif import_type == 'benefit_bulk':
            year = request.data.get('year') or request.query_params.get('year')
            month = request.data.get('month') or request.query_params.get('month')
            year = int(year) if year else None
            month = str(month) if month else None
            created, skipped = ImportEngine.import_benefits(valid_rows, company, year=year, month=month)
        elif import_type == 'attendance_bulk':
            created, skipped = ImportEngine.import_attendance(valid_rows, company)
        elif import_type == 'leave_bulk':
            created, skipped = ImportEngine.import_leaves(valid_rows, company)
        else:
            created, skipped = ImportEngine.import_simple(valid_rows, company, import_type)
    except Exception as e:
        return Response({'error': f'خطا در درون‌ریزی: {str(e)[:150]}'}, status=500)

    return Response({
        'message': 'درون‌ریزی با موفقیت انجام شد',
        'import_type': import_type,
        'import_type_label': ImportEngine.IMPORT_TYPES[import_type]['label'],
        'total_rows': len(rows),
        'valid_rows': len(valid_rows),
        'imported_count': created,
        'skipped_count': len(skipped),
        'skipped': skipped[:50],
    })